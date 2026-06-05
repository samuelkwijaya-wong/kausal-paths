#!/usr/bin/env python
"""
Build the King County GHG inventory input dataset from the source spreadsheet.

Reads the ``All`` tab of the King County GHG trend spreadsheet (the row-level
inventory database) and produces a CSV in the standard ``upload_new_dataset``
wide format (see ``docs/data-management.md``), with one column per inventory
year and one row per (subsector x breakdown-dimension) combination.

The emission **subsectors** follow the ``Summary - County`` tab layout, which
differs from the raw ``Activity_Name`` values in the ``All`` tab:

  * Built Environment "Other sources" rows are split into ``fuel_oil`` /
    ``propane`` by ``Activity_FuelType``.
  * "Industrial process" / "Industrial Process" (casing variants) are merged
    into ``industrial_processes``.
  * Land Use "Tree loss" -> ``forest_and_trees``; "Tree sequestration" ->
    ``forest_and_tree_sequestration`` (negative); Solid Waste "Solid waste
    disposal sequestration" -> ``solid_waste_disposal_sequestration``
    (negative).

The breakdown **dimensions** come from ``All`` columns H-O. Each subsector
node in ``configs/king-county.yaml`` flattens (sums over) these dimensions, so
they are carried in the data for future drill-down but summed away today.

2007 is excluded (it is an exception, and is absent from the ``All`` tab
anyway).

Usage::

    python tools/build_king_county_dataset.py \
        --xlsx data/king-county/King_County_GHG_Trend_2007_2023.xlsx \
        --out  data/king-county/ghg_inventory.csv

Then upload to DVC and wire up the config (see docs/data-management.md)::

    python -m notebooks.upload_new_dataset \
        --input-csv data/king-county/ghg_inventory.csv \
        --output-dvc king-county --language en --instance king-county
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import polars as pl

# --- Source layout (All tab) ----------------------------------------------
# Header is on row 4, data from row 5. Columns (1-indexed):
#   C(3) Year, D(4) Activity_Type, E(5) Activity_Name,
#   H(8) Activity_Utility, I(9) Activity_Sector, J(10) Activity_VehicleType,
#   K(11) Activity_MOVESSector, L(12) Activity_FuelType, M(13) Activity_Livestock,
#   N(14) Activity_ReportDetail, O(15) Activity_MaterialType, X(24) MTCO2e
COL_YEAR = 3
COL_TYPE = 4
COL_NAME = 5
COL_JURISDICTION = 6  # column F
COL_FUEL_TYPE = 12  # column L (Activity_FuelType)
COL_MTCO2E = 24

# Community-wide county inventory uses only the county-level rows. The All tab
# also contains rows for individual cities and neighbouring counties (used for
# the per-city inventories on the 'Summary - Cities' tab); those must be excluded.
JURISDICTION = 'King County'
# H-O breakdown dimensions -> model dimension id
DIM_COLS = {
    8: 'utility',
    9: 'economic_sector',
    10: 'vehicle_type',
    11: 'moves_sector',
    12: 'fuel_type',
    13: 'livestock',
    14: 'report_detail',
    15: 'material_type',
}
DIM_IDS = list(DIM_COLS.values())

EXCLUDED_YEARS = {2007}

DATASET_NAME = 'ghg_inventory'
EMISSION_UNIT = 't_co2e/a'


def slugify(value: object) -> str | None:
    """
    Lowercase, collapse non-alphanumerics to single underscores.

    Returns None for blank / not-applicable cells (None, '', or the literal
    0 placeholder used in Activity_Utility).
    """
    if value is None:
        return None
    if isinstance(value, (int, float)) and value == 0:
        return None
    s = str(value).strip()
    if not s or s == '0':
        return None
    s = s.lower()
    s = re.sub(r'[^a-z0-9]+', '_', s)
    return s.strip('_') or None


# --- Subsector mapping (Summary - County tab) ------------------------------
# Activity_Name -> subsector id, keyed by Activity_Type. "Industrial process" /
# "Industrial Process" are casing variants merged into one subsector.
_BY_TYPE: dict[str, dict[str, str]] = {
    'Built Environment': {
        'Electricity': 'electricity',
        'Natural gas': 'natural_gas',
        'Industrial process': 'industrial_processes',
        'Industrial Process': 'industrial_processes',
    },
    'Transportation & Other Mobile Sources': {
        'On-road vehicles': 'on_road_vehicles',
        'Aviation': 'aviation',
        'Off-road equipment': 'off_road_equipment',
        'Marine vessels and rail': 'marine_vessels_and_rail',
    },
    'Solid Waste & Wastewater': {
        'Solid waste generation and disposal': 'solid_waste_generation_and_disposal',
        'Wastewater process emissions': 'wastewater_process_emissions',
        'Solid waste disposal sequestration': 'solid_waste_disposal_sequestration',
    },
    'Refrigerants': {'Refrigerants': 'refrigerants'},
    'Land Use': {
        'Agriculture': 'agriculture',
        'Tree loss': 'forest_and_trees',
        'Tree sequestration': 'forest_and_tree_sequestration',
    },
}
# Built Environment "Other sources" rows are split by fuel type (Summary lists
# Fuel oil and Propane as separate subsectors).
_OTHER_SOURCES_BY_FUEL = {'fuel oil': 'fuel_oil', 'propane': 'propane'}


def subsector_of(atype: str | None, aname: str | None, fuel_type: object) -> str | None:
    """Map an All-tab (Activity_Type, Activity_Name) pair to a Summary subsector id."""
    if atype == 'Built Environment' and aname == 'Other sources':
        ft = str(fuel_type).strip().lower() if fuel_type is not None else ''
        return _OTHER_SOURCES_BY_FUEL.get(ft)
    return _BY_TYPE.get(atype or '', {}).get(aname or '')


def normalize_dim(dim_id: str, value: object) -> str | None:
    """Slugify a breakdown-dimension cell, merging known casing variants."""
    slug = slugify(value)
    if slug is None:
        return None
    if dim_id == 'fuel_type' and slug == 'jet_fuel':  # 'Jet Fuel' / 'Jet fuel'
        return 'jet_fuel'
    return slug


def read_rows(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb['All']
    records: list[dict] = []
    skipped: dict[tuple, int] = {}
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        atype = row[COL_TYPE - 1]
        aname = row[COL_NAME - 1]
        if atype is None and aname is None:
            continue
        if row[COL_JURISDICTION - 1] != JURISDICTION:
            continue
        year = row[COL_YEAR - 1]
        if year is None or int(year) in EXCLUDED_YEARS:
            continue
        fuel_type = row[COL_FUEL_TYPE - 1]  # Activity_FuelType (column L)
        sub = subsector_of(atype, aname, fuel_type)
        if sub is None:
            skipped[(atype, aname)] = skipped.get((atype, aname), 0) + 1
            continue
        mt = row[COL_MTCO2E - 1]
        rec: dict = {'subsector': sub, 'Year': int(year), 'Value': float(mt or 0.0)}
        for col, dim_id in DIM_COLS.items():
            rec[dim_id] = normalize_dim(dim_id, row[col - 1])
        records.append(rec)
    if skipped:
        print('WARNING: skipped unmapped (Activity_Type, Activity_Name) rows:', file=sys.stderr)
        for key, n in sorted(skipped.items(), key=lambda x: str(x[0])):
            print(f'  {key}: {n} rows', file=sys.stderr)
    return records


def build(records: list[dict]) -> pl.DataFrame:
    df = pl.DataFrame(records)
    group_cols = ['subsector', *DIM_IDS, 'Year']
    agg = df.group_by(group_cols).agg(pl.col('Value').sum())
    # Long -> wide on Year
    wide = agg.pivot(values='Value', index=['subsector', *DIM_IDS], on='Year', aggregate_function='sum')
    year_cols = sorted(int(c) for c in wide.columns if c not in ('subsector', *DIM_IDS))
    wide = wide.select(['subsector', *DIM_IDS, *[str(y) for y in year_cols]])
    # Standard upload metadata columns
    wide = wide.with_columns(
        pl.lit('Value').alias('Metric'),
        pl.lit(EMISSION_UNIT).alias('Unit'),
        pl.lit('').alias('Quantity'),
        pl.lit(DATASET_NAME).alias('Dataset'),
    )
    ordered = ['Metric', 'Unit', 'Quantity', 'Dataset', 'subsector', *DIM_IDS, *[str(y) for y in year_cols]]
    return wide.select(ordered).sort(['subsector', *DIM_IDS])


def dump_dimension_categories(records: list[dict]) -> None:
    """Print distinct slug ids + an example original label per dimension, to fill the YAML."""
    print('\n=== Dimension categories (id only; fill labels in YAML) ===')
    subs = sorted({r['subsector'] for r in records})
    print(f'subsector ({len(subs)}): {subs}')
    for dim_id in DIM_IDS:
        vals = sorted({r[dim_id] for r in records if r[dim_id] is not None})
        print(f'{dim_id} ({len(vals)}): {vals}')


def validate(df: pl.DataFrame, records: list[dict]) -> None:
    """Print per-subsector yearly totals so they can be eyeballed vs the Summary tab."""
    print('\n=== Per-subsector totals by year (compare to Summary - County) ===')
    year_cols = [c for c in df.columns if c.isdigit()]
    tot = df.group_by('subsector').agg([pl.col(y).sum().alias(y) for y in year_cols]).sort('subsector')
    with pl.Config(tbl_rows=30, tbl_cols=20, fmt_str_lengths=40, float_precision=0):
        print(tot)
    grand = {y: df[y].sum() for y in year_cols}
    print('\nGrand total (gross + sequestration) by year:')
    for y in year_cols:
        print(f'  {y}: {grand[y]:,.0f}')


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--xlsx', type=Path, default=Path('data/king-county/King_County_GHG_Trend_2007_2023.xlsx'))
    ap.add_argument('--out', type=Path, default=Path('data/king-county/ghg_inventory.csv'))
    ap.add_argument('--dump-dims', action='store_true', help='Print distinct dimension category ids and exit')
    args = ap.parse_args()

    records = read_rows(args.xlsx)
    print(f'Read {len(records)} mapped records.')

    if args.dump_dims:
        dump_dimension_categories(records)
        return

    df = build(records)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    df.write_csv(args.out, null_value='')
    print(f'Wrote {df.height} rows x {df.width} cols -> {args.out}')
    dump_dimension_categories(records)
    validate(df, records)


if __name__ == '__main__':
    main()
